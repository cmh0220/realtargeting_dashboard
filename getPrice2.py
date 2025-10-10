# 아파트 정보 조회 후 엑셀저장 (카페 지원용)

import json
import time
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill


################################## 컬럼 자동정렬
def AutoFitColumnSize(worksheet, columns=None, margin=5):
    for i, column_cells in enumerate(worksheet.columns):
        is_ok = False
        if columns == None:
            is_ok = True
        elif isinstance(columns, list) and i in columns:
            is_ok = True

        if is_ok:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin

    return worksheet

box = Border(left=Side(border_style='thin', color = 'FF000000'),
             right=Side(border_style='thin', color = 'FF000000'),
             top=Side(border_style='thin', color = 'FF000000'),
             bottom=Side(border_style='thin', color = 'FF000000'))


################################## 날짜변수 선언
now = datetime.now()
formattedDate = now.strftime("%Y%m%d")


################################## 아파트정보 및 요청자 정보 입력
apt_name = [
      '공작럭키'
    ]
comp = [
      '1460'
        ]
wb_total = Workbook()  # create xlsx file
ws_total = wb_total.active  # create xlsx sheet
ws_total.append(["단지명", "Comp", "일자", "매매개수", "전세개수"])
#################### 한싸이클 시작

for i in range(len(apt_name)):

    # 엑셀파일명 변수
    save_file_name =  apt_name[i] + '_' + formattedDate + '_' + comp[i]
    # 매물리스트 변수
    list_sum = []
    # 엑셀워크북 변수
    wb = Workbook()  # create xlsx file
    ws = wb.active  # create xlsx sheet


    def SaveAptInfo(p_page, p_comp, p_list_sum):
        down_url = 'https://new.land.naver.com/api/articles/complex/' + p_comp + '?realEstateType=APT&tradeType=A1%3AB1&tag=%3A%3A%3A%3A%3A%3A%3A%3A&rentPriceMin=0&rentPriceMax=900000000&priceMin=0&priceMax=900000000&areaMin=0&areaMax=900000000&oldBuildYears&recentlyBuildYears&minHouseHoldCount&maxHouseHoldCount&showArticle=false&sameAddressGroup=true&minMaintenanceCost&maxMaintenanceCost&priceType=RETAIL&directions=&page=' + p_page + '&complexNo=' + p_comp + '&buildingNos=&areaNos=&type=list&order=rank'
        r = requests.get(down_url, data={"sameAddressGroup": "true"}, headers={
            "Accept-Encoding": "gzip",
            "Host": "m.land.naver.com",
            "Referer": "https://new.land.naver.com/",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin",
            "User-Agent": "Mozilla/5.0 (Phone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.3 Mobile/15E148 Safari/604.1",
            "authorization": "Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6IlJFQUxFU1RBVEUiLCJpYXQiOjE2NDk4MzA1OTgsImV4cCI6MTY0OTg0MTM5OH0.lNapCTDOQAQZwIGafsVQeUqBmSzUTKm8dUJHaJJgrK0"
        })

        r.encoding = "utf-8-sig"

        temp = json.loads(r.text)

        p_apt_list = temp['articleList']

        # 추출한 정보를 리스트에 적재
        for data in p_apt_list:

            # 추출한 정보 로그
            # print(data)

            if data.get("articleFeatureDesc") == None:

                # 금액 숫자로 변환
                dowp = str(data['dealOrWarrantPrc'])
                dowp_int = int(dowp.replace(',', '')) if dowp.find('억') == -1 else (int(dowp[
                                                                                        :dowp.find('억')]) * 10000) + (
                                                                                       0 if dowp.endswith('억') else int(
                                                                                           dowp[
                                                                                           dowp.find('억') + 1:].replace(
                                                                                               ',', '')))

                p_list_sum.append(
                    [data['articleNo'], data['articleName'], data['tradeTypeName'], data['area1'],
                     round(float(data['area1']) / 3.3), data['area2'], round(float(data['area2']) / 3.3),
                     data['buildingName'], data['floorInfo'], data['direction'], dowp_int,
                     round(dowp_int / round(float(data['area1']) / 3.3)),
                     '', data['realtorName']])
            else:

                # 금액 숫자로 변환
                dowp = str(data['dealOrWarrantPrc'])
                dowp_int = int(dowp.replace(',', '')) if dowp.find('억') == -1 else (int(dowp[
                                                                                        :dowp.find('억')]) * 10000) + (
                                                                                       0 if dowp.endswith('억') else int(
                                                                                           dowp[
                                                                                           dowp.find('억') + 1:].replace(
                                                                                               ',', '')))

                p_list_sum.append(
                    [data['articleNo'], data['articleName'], data['tradeTypeName'], data['area1'],
                     round(float(data['area1']) / 3.3), data['area2'], round(float(data['area2']) / 3.3),
                     data['buildingName'], data['floorInfo'], data['direction'], dowp_int,
                     round(dowp_int / round(float(data['area1']) / 3.3)),
                     data['articleFeatureDesc'], data['realtorName']])
        return p_list_sum, p_apt_list


    # 1회차 조회
    list_sum, apt_list = SaveAptInfo('1', comp[i], list_sum)

    # 2회차 조회
    if(len(apt_list) == 20):
        list_sum, apt_list = SaveAptInfo('2', comp[i], list_sum)

    # 3회차 조회
    if(len(apt_list) == 20):
        list_sum, apt_list = SaveAptInfo('3', comp[i], list_sum)

    # 4회차 조회
    if(len(apt_list) == 20):
        list_sum, apt_list = SaveAptInfo('4', comp[i], list_sum)

    # 5회차 조회
    if(len(apt_list) == 20):
        list_sum, apt_list = SaveAptInfo('5', comp[i], list_sum)

    # 6회차 조회
    if(len(apt_list) == 20):
        list_sum, apt_list = SaveAptInfo('6', comp[i], list_sum)

    # 7회차 조회
    if(len(apt_list) == 20):
        list_sum, apt_list = SaveAptInfo('7', comp[i], list_sum)

    # 정렬
    list_sum.sort(key=lambda x : (x[2], x[3], str(x[10])))

    #매매, 전세 물건개수 카운팅
    m_cnt = 0
    j_cnt = 0
    for ll in list_sum:
        if ll[2] == '매매':
            m_cnt += 1
        elif ll[2] == '전세':
            j_cnt += 1


    ws.append(['물건번호', '단지명', '구분', '공급', '평', '전용', '평', '동', '층', '향', '가격', '평당가', '매물특징', '공인중개사'])  # input header (first row)

    for ll in list_sum:
        ws.append(ll)

    #셀 가시성 개선
    for row in ws.rows:
        for cell in row:

            #가운데정렬
            cell.alignment = Alignment(horizontal = "center", vertical = "center")

            #색 구분
            if cell.value == "매매":
                cell.fill = PatternFill(fgColor = "95B3D7", fill_type="solid")

            elif cell.value == "전세":
                cell.fill = PatternFill(fgColor="FABF8F", fill_type="solid")

            # 셀 크기 자동조절#

    AutoFitColumnSize(ws, margin = 5)

    for num, row in enumerate(ws.rows):
        for cell in row:
                cell.border = box

    wb.save('C:/Users/notre/OneDrive/문서/네이버매물/' + save_file_name + '.xlsx')

    print(apt_name[i] + '(' + comp[i] + ') ' + formattedDate + ' '+ str(m_cnt).zfill(3) + ' ' + str(j_cnt).zfill(3))

    ws_total.append([apt_name[i], comp[i], formattedDate, str(m_cnt), str(j_cnt)])

    ######################한싸이클 완료
    time.sleep(1)

####################### 아파트리스트 전체 매물개수 저장
wb_total.save('C:/Users/notre/OneDrive/문서/네이버매물/' + formattedDate + '_전체.xlsx')
print(formattedDate + ' 처리완료!!')